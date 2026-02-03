from motor.motor_asyncio import AsyncIOMotorDatabase
from bson import ObjectId
from typing import List, Optional, Dict, Any
from io import BytesIO
import asyncio
from dataclasses import fields

from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML
from starlette.datastructures import UploadFile

from app.shared.storage.s3.objects import storage_s3_retrieve_objects_url, generate_s3_temporary_storage_object_key
from app.shared.storage.utils import download_file_base64
from app.shared.datetime import time_now
import pandas
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from pathlib import Path

from colorsys import hls_to_rgb
from app.modules.report.report_choices import HierarchyStandChoice, ImageExportModeChoice
from app.shared.database.pipelines.inventory_items import InventoryItemsPipelines

from app.shared.stream.image_zipstream import ImageStreamingZipWriter
from app.shared.storage.s3.multi_part_uploader import MultipartUploader
from app.shared.global_functions.download_storage_objects import DownloadStorageObjecs
import uuid
from app.shared.files.files_type_choices import FileTypeChoices


from app.modules.report.report_schemas import (
    InventoryResposabilityAgreementItemDTO,
    InventoryResposabilityAgreementLocationDTO,
    AnalyticalReportRawDataDTO
)


class AssetInventoryResponsibilityReportService:
    def __init__(self, database: AsyncIOMotorDatabase):
        self.db = database

    async def create_inventory_responsibility_agreement_report(
        self,
        parent_location_ids: Optional[List[str]]
    ) -> UploadFile:

        data = await self._build_data(parent_location_ids)
        env = Environment(
            loader=FileSystemLoader("app/template"),
            autoescape=True
        )
        template = env.get_template(
            "report/inventory_responsibility_agreement/report.html"
        )
        html_content = template.render(
            data=data,
            generated_at=time_now().strftime("%d/%m/%Y %H:%M")
        )

        pdf_bytes = BytesIO()
        HTML(string=html_content).write_pdf(pdf_bytes)
        pdf_bytes.seek(0)

        return UploadFile(
            filename="termo_responsabilidade_inventario.pdf",
            file=BytesIO(pdf_bytes.read())
        )

    async def _build_data(
        self,
        parent_location_ids: Optional[List[str]]
    ) -> List[InventoryResposabilityAgreementLocationDTO]:

        locations_ids = await self._get_all_descendant_locations(parent_location_ids)
        item_docs = await self.db.inventory_items.aggregate([
            {
                "$match": {
                    "parent_id": {"$in": locations_ids},
                    "node_type": {"$ne": "LOCATION"}
                }
            },
            {
                "$graphLookup": {
                    "from": "inventory_items",
                    "startWith": "$_id",
                    "connectFromField": "_id",
                    "connectToField": "parent_id",
                    "as": "descendants",
                    "restrictSearchWithMatch": {"node_type": {"$ne": "LOCATION"}}
                }
            },
            {"$addFields": {"root_loc": "$parent_id"}},
            {"$project": {"docs": {"$concatArrays": [["$$ROOT"], "$descendants"]}, "root_loc": 1}},
            {"$unwind": "$docs"},
            {"$addFields": {"docs.root_loc": "$root_loc"}},
            {"$replaceRoot": {"newRoot": "$docs"}},
            {"$project": {"descendants": 0}}
        ]).to_list(None)

        location_map = {
            location_doc["_id"]: InventoryResposabilityAgreementLocationDTO(
                path=" -> ".join(location_doc.get("path", [])),
                reference=location_doc.get("reference", None),
                level=location_doc.get("level", None)
            )
            for location_doc in (
                await self.db.inventory_items.find(
                    {"_id": {"$in": locations_ids}},
                    {"_id": 1, "path": 1, "reference": 1, "level": 1}
                ).to_list(None)
            )
        }

        item_map = {
            item_doc["_id"]: item_doc 
            for item_doc in item_docs
        }

        for item_doc in item_docs:
            parent = None
            parent_id = item_doc.get("parent_id")

            if parent_id and parent_id in item_map:
                parent = item_map[parent_id]

            location = location_map.get(item_doc.get("root_loc"))
            if not location:
                continue

            level = (
                item_doc.get("level") - location.level
                if parent
                else 1
            )

            photo_key = item_doc.get("photos")[0] if item_doc.get("photos", None) else None
            location.items.append(
                InventoryResposabilityAgreementItemDTO(
                    reference=item_doc.get("reference", None),
                    checked=item_doc.get("checked", False),
                    description=item_doc.get("asset_data", {}).get("description"),
                    serial=item_doc.get("asset_data", {}).get("serial"),
                    model=item_doc.get("asset_data", {}).get("model"),
                    color=self.level_to_color(level),
                    photo_key=photo_key,
                    parent_reference=parent.get("reference") if parent else None                
                )
            )

        for location in location_map.values():
            location.items = self._sort_items_hierarchy(location.items)

        sorted_locations = sorted(
            location_map.values(),
            key=lambda loc: len(loc.items) == 0
        )

        await self.download_by_path(sorted_locations)
        return list(sorted_locations)

    async def download_by_path(
        self,
        locations: List[InventoryResposabilityAgreementLocationDTO]
    ) -> None:

        semaphore = asyncio.Semaphore(6)

        async def resolve(item: InventoryResposabilityAgreementItemDTO):
            if not item.photo_key:
                return
            async with semaphore:
                try:
                    url = await storage_s3_retrieve_objects_url(item.photo_key)
                    item.photo_base64 = await download_file_base64(url)
                except Exception:
                    item.photo_base64 = None

        tasks = [
            resolve(item)
            for location in locations
            for item in location.items
            if item.photo_key
        ]

        if tasks:
            await asyncio.gather(*tasks)

    async def _get_all_descendant_locations(
        self,
        parent_location_ids: Optional[List[str]]
    ) -> List[ObjectId]:

        if not parent_location_ids:
            docs = await self.db.inventory_items.find(
                {"node_type": "LOCATION"}, 
                {"_id": 1}
            ).to_list(None)

            return [doc["_id"] for doc in docs]

        parent_location_object_ids = []
        for parent_location_id in parent_location_ids:
            if not ObjectId.is_valid(parent_location_id):
                raise ValueError(
                    f"Invalid parent_location_id sent '{parent_location_id}', can not be tranformed to ObjectId"
                )
            parent_location_object_ids.append(ObjectId(parent_location_id))

        docs = await self.db.inventory_items.aggregate([
            {"$match": {"_id": {"$in": parent_location_object_ids}}},
            {
                "$graphLookup": {
                    "from": "inventory_items",
                    "startWith": "$_id",
                    "connectFromField": "_id",
                    "connectToField": "parent_id",
                    "as": "descendants"
                }
            },
            {"$unwind": "$descendants"},
            {"$match": {"descendants.node_type": "LOCATION"}},
            {"$group": {"_id": "$descendants._id"}},
            {"$project": {"_id": 1}}
        ]).to_list(None)

        descendant_locations_ids = [doc["_id"] for doc in docs]
        return list(set(descendant_locations_ids + parent_location_object_ids))

    def level_to_color(self, level: int) -> str:
        if level == 1:
            return "#f8f9fa"

        hue = (level * 0.25) % 1.0  

        lightness = 0.95 - (level * 0.02)
        lightness = max(0.75, lightness)

        saturation = 0.35 if level <= 5 else 0.3

        r, g, b = hls_to_rgb(hue, lightness, saturation)
        color = "#{:02X}{:02X}{:02X}".format(int(r*255), int(g*255), int(b*255))
        return color

    def _sort_items_hierarchy(self, items: List[InventoryResposabilityAgreementItemDTO]) -> List[InventoryResposabilityAgreementItemDTO]:
        children_map: Dict[Optional[str], List[InventoryResposabilityAgreementItemDTO]] = {}
        for item in items:
            parent = item.parent_reference
            children_map.setdefault(parent, []).append(item)

        parents_set = {item.parent_reference for item in items if item.parent_reference is not None}

        def flatten(parent_id: Optional[str]):
            flat_list = []

            leaves = [item for item in children_map.get(parent_id, []) if item.reference not in parents_set]
            for leaf in leaves:
                flat_list.append(leaf)

            non_leaves = [item for item in children_map.get(parent_id, []) if item.reference in parents_set]
            for parent_item in non_leaves:
                flat_list.append(parent_item)
                flat_list.extend(flatten(parent_item.reference))

            return flat_list

        return flatten(None)

class AnalyticalReportService:
    def __init__(self, database: AsyncIOMotorDatabase):
        self.db = database

    async def create_analytical_report(
        self,
        parent_ids: Optional[List[str]]
    ) -> UploadFile:

        rows = await self._build_data(parent_ids)
        dataframe = pandas.DataFrame(rows)
        file_bytes = BytesIO()

        with pandas.ExcelWriter(file_bytes, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=False, sheet_name='Relatório')
            ws = writer.sheets['Relatório']

            header_fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for col_num, column_title in enumerate(dataframe.columns, 1):
                cell = ws[f'{get_column_letter(col_num)}1']
                cell.fill = header_fill
                cell.font = header_font

            for col_num, column in enumerate(dataframe.columns, 1):
                max_length = max(
                    dataframe[column].astype(str).map(len).max(),
                    len(column)
                )
                ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        file_bytes.seek(0)

        return UploadFile(
            filename="relatorio_analitico.xlsx",
            file=BytesIO(file_bytes.read())
        )

    async def _build_data(
        self,
        parent_ids: Optional[List[str]]
    ) -> List[Dict[str, Any]]:

        raw_data = await self._get_raw_data(parent_ids)
        max_loc_path_length = 0
        max_hierarchy_path_length = 0

        for item in raw_data:
            if item.location_path:
                max_loc_path_length = max(max_loc_path_length, len(item.location_path))
            if item.hierarchy_path:
                max_hierarchy_path_length = max(max_hierarchy_path_length, len(item.hierarchy_path))

        rows: List[Dict[str, Any]] = []
        for item in raw_data:
            row: Dict[str, Any] = {}

            for index in range(max_loc_path_length):
                col_name = f"LOC {index+1}"
                path_value = None
                if item.location_path and index < len(item.location_path):
                    path_value = item.location_path[index]
                row[col_name] = path_value

            row["ATIVO"] = item.reference
            row["ATIVO PRINCIPAL"] = item.hierarchy_path[0]
            row["NIVEL"] = f"Nivel {item.level}"

            for index in range(max_hierarchy_path_length):
                col_name = f"NIVEL {index+1}"
                path_value = None
                if item.hierarchy_path and index < len(item.hierarchy_path):
                    path_value = item.hierarchy_path[index]
                row[col_name] = path_value

            row["HIERARQUIA"] = (
                "Filho" 
                if item.hierarchy_stand == HierarchyStandChoice.CHILD 
                else "Pai"
            )

            row["DESCRIÇÃO"] = item.asset_data.get("description")
            row["FABRICANTE"] = item.asset_data.get("manufacturer")
            row["MODELO"] = item.asset_data.get("model")
            row["N ID TEC"] = item.asset_data.get("n_id_tec")
            row["TIPO INTERNO"] = item.asset_data.get("type")
            row["TECNOLOGIA"] = None
            row["STATUS OPERACIONAL"] = None
            row["PROJETO"] = None
            row["HOSTNAME"] = item.asset_data.get("hostname")
            row["CTG EQUIPAMENTO"] = None
            row["EQUIPAMENTO CONFIRMADO"] = None
            row["STATUS EM CAMPO"] = None
            row["SISTEMA DE GERÊNCIA"] = None
            row["STATUS DA CONCILIAÇÃO COM A CONTABILIDADE"] = None
            row["COM INVENTÁRIO EM CAMPO?"] = None
            row["INVENTARIANTE"] = None
            row["STATUS DA CONCILIAÇÃO COM A GERÊNCIA"] = None
            row["LOCALIZAÇÃO"] = None
            row["MÊS PREV"] = None
            row["STATUS CAMPO V.TAL"] = None
            row["LOCALIZAÇÃO_2"] = None
            row["INVENTARIADO"] = "SIM" if item.checked else "NÃO"
            row["INVENTARIADO EM"] = item.checked_at.strftime("%d/%m/%Y %H:%M") if item.checked_at else None
            rows.append(row)

        return rows        

    async def _get_raw_data(
        self,
        parent_ids: Optional[List[str]]
    ) -> List[AnalyticalReportRawDataDTO]:

        docs = []
        raw_fields = {
            field.name: 1
            for field in fields(AnalyticalReportRawDataDTO)
        }

        parent_object_ids = []

        if not parent_ids:
            docs = await self.db.inventory_items.find(
                {"node_type": "ASSET"},
                raw_fields
            ).to_list(None)
            parent_object_ids = [doc["_id"] for doc in docs]
        else:
            for parent_id in parent_ids:
                if not ObjectId.is_valid(parent_id):
                    raise ValueError(
                        f"Invalid parent_id sent '{parent_id}', can not be tranformed to ObjectId"
                    )
                parent_object_ids.append(ObjectId(parent_id))

        docs = await self.db.inventory_items.aggregate(
        [
            {
                "$match": {
                    "_id": { "$in": parent_object_ids }
                }
            },
            {
                "$graphLookup": {
                    "from": "inventory_items",
                    "startWith": "$_id",
                    "connectFromField": "_id",
                    "connectToField": "parent_id",
                    "as": "descendants"
                }
            },
            {
                "$project": {
                    "items": {
                        "$concatArrays": [
                            ["$$ROOT"],
                            "$descendants"
                        ]
                    }
                }
            },
            { "$unwind": "$items" },
            { "$replaceRoot": { "newRoot": "$items" } },
            {
                "$group": {
                    "_id": "$_id",
                    "doc": { "$first": "$$ROOT" }
                }
            },
            {
                "$replaceRoot": {
                    "newRoot": "$doc"
                }
            },
            {
                "$match": {
                    "node_type": "ASSET"
                }
            },
            {
                "$graphLookup": {
                    "from": "inventory_items",
                    "startWith": "$parent_id",
                    "connectFromField": "parent_id",
                    "connectToField": "_id",
                    "as": "ancestors",
                    "depthField": "depth"
                }
            },
            {
                "$addFields": {
                    "ancestors": {
                        "$sortArray": {
                            "input": "$ancestors",
                            "sortBy": { "depth": -1 }
                        }
                    }
                }
            },
            {
                "$addFields": {
                    "location_path": {
                        "$cond": [
                            { "$eq": ["$node_type", "LOCATION"] },
                            {
                                "$concatArrays": [
                                    {
                                        "$map": {
                                            "input": {
                                                "$filter": {
                                                    "input": "$ancestors",
                                                    "as": "a",
                                                    "cond": { "$eq": ["$$a.node_type", "LOCATION"] }
                                                }
                                            },
                                            "as": "loc",
                                            "in": "$$loc.reference"
                                        }
                                    },
                                    ["$reference"]
                                ]
                            },
                            {
                                "$map": {
                                    "input": {
                                        "$filter": {
                                            "input": "$ancestors",
                                            "as": "a",
                                            "cond": { "$eq": ["$$a.node_type", "LOCATION"] }
                                        }
                                    },
                                    "as": "loc",
                                    "in": "$$loc.reference"
                                }
                            }
                        ]
                    },
                    "hierarchy_path": {
                        "$concatArrays": [
                            {
                                "$map": {
                                    "input": {
                                        "$filter": {
                                            "input": "$ancestors",
                                            "as": "a",
                                            "cond": { "$eq": ["$$a.node_type", "ASSET"] }
                                        }
                                    },
                                    "as": "asset",
                                    "in": "$$asset.reference"
                                }
                            },
                            ["$reference"]
                        ]
                    }
                }
            },
            {
                "$addFields": {
                    "level": {
                        "$subtract": [
                            {
                                "$add": [
                                    "$level",
                                    1
                                ]
                            },
                            { "$size": "$location_path" }
                        ]
                    }
                }
            },
            {
                "$addFields": {
                    "hierarchy_stand": {
                        "$cond": [
                            { "$eq": ["$level", 1] },
                            HierarchyStandChoice.PARENT.value,
                            HierarchyStandChoice.CHILD.value
                        ]
                    }
                }
            },
            {
                "$project": raw_fields
            },
            {
                "$sort": {
                    "level": 1
                }
            }
        ]
        ).to_list(None)

        print("len", len(docs))

        dto_list: List[AnalyticalReportRawDataDTO] = [
            AnalyticalReportRawDataDTO(**doc) for doc in docs
        ]

        return dto_list

class ImagesExportService:
    def __init__(self, database: AsyncIOMotorDatabase):
        self.db = database

    async def export_images(self, parent_id: Optional[str], mode: ImageExportModeChoice):
        storage_key = None
        if mode == ImageExportModeChoice.EXPORT_ALL:
            storage_key = await self.export_all_images()
        elif mode == ImageExportModeChoice.EXPORT_SINGLE:
            if not parent_id:
                raise ValueError(
                    f"parent_id inválido: valores vazios não são aceitos para a operação EXPORT_SINGLE."
                )
            storage_key = await self.export_single_item_images(parent_id)
        else:
            if not parent_id:
                raise ValueError(
                    f"parent_id inválido: valores vazios não são aceitos para a operação EXPORT_TREE."
                )
            storage_key = await self.export_tree_items_images(parent_id)
        return storage_key

    async def export_all_images(self):
        pipeline_service = InventoryItemsPipelines(self.db)
        all_location_docs = await pipeline_service.get_all_locations_with_parent_path(
            projection_fields={"_id": 1, "parent_locations": 1}
        )

        locations_map: Dict[str, str] = {
            str(loc["_id"]): " -> ".join(loc.get("parent_locations", ["POSICAO_SEM_PATH"]))
            for loc in all_location_docs
        }

        items_cursor = await pipeline_service.get_all_items_by_locations(
            locations_ids=[loc["_id"] for loc in all_location_docs],
            projection_fields={"root_loc": 1, "reference": 1, "photos": 1},
            batch_size=200,
            as_list=False
        )

        storage_key = generate_s3_temporary_storage_object_key(FileTypeChoices.ZIP)
        uploader = MultipartUploader(
            key=storage_key
        )

        zip_writer = ImageStreamingZipWriter(uploader)

        async for item in items_cursor:
            root_loc = str(item["root_loc"])
            folder = locations_map.get(root_loc)

            if not folder:
                continue

            photos_keys = item.get("photos") or []
            if not photos_keys:
                continue

            photos_base64 = await DownloadStorageObjecs().download_by_path(photos_keys)

            await zip_writer.process(
                folder=folder,
                reference=item.get("reference", "SEM_REFERENCIA"),
                images_base64=photos_base64
            )

        await zip_writer.stream_to_cloud()
        return storage_key

    async def export_single_item_images(self, asset_id: str):
        if not ObjectId.is_valid(asset_id):
            raise ValueError(
                f"ID de asset inválido: {asset_id} não pode ser convertido para ObjectId."
            )

        storage_key = generate_s3_temporary_storage_object_key(
            FileTypeChoices.ZIP
        )
        uploader = MultipartUploader(
            key=storage_key
        )
        zip_writer = ImageStreamingZipWriter(
            uploader
        )

        asset_object_id = ObjectId(asset_id)
        pipeline_service = InventoryItemsPipelines(self.db)
        item = await pipeline_service.get_asset_with_images_and_parent_locations(
            asset_id=asset_object_id,
            projection_fields={"locations": "$locations.reference", "reference": 1, "photos": 1}
        )

        if item:
            location_path = (" -> ").join(item.get("locations", ["CAMINHO_LOCALIZACAO_NAO_ENCONTRADO"]))
            photos_keys = item.get("photos") or []
            photos_base64 = await DownloadStorageObjecs().download_by_path(photos_keys)

            await zip_writer.process(
                folder=location_path,
                reference=item.get("reference", "SEM_REFERENCIA"),
                images_base64=photos_base64
            )

        await zip_writer.stream_to_cloud()
        return storage_key

    async def export_tree_items_images(self, parent_id: str):
        if not ObjectId.is_valid(parent_id):
            raise ValueError(
                f"ID de asset inválido: {parent_id} não pode ser convertido para ObjectId."
            )

        storage_key = generate_s3_temporary_storage_object_key(
            FileTypeChoices.ZIP
        )
        uploader = MultipartUploader(
            key=storage_key
        )
        zip_writer = ImageStreamingZipWriter(
            uploader
        )

        pipeline_service = InventoryItemsPipelines(self.db)
        parent_object_id = ObjectId(parent_id)

        items_cursor = await pipeline_service.get_asset_tree_with_images_and_parent_locations(
            parent_id=parent_object_id,
            projection_fields={"parent_locations": "$parent_locations.reference", "reference": 1, "photos": 1},
            batch_size=200,
            as_list=False
        )

        async for item in items_cursor:
            locations_path = item.get("parent_locations")
            if not locations_path:
                locations_path = ["CAMINHO_LOCALIZACAO_NAO_ENCONTRADO"]

            folder = " -> ".join(locations_path)
            photos_keys = item.get("photos", [])
            photos_base64 = await DownloadStorageObjecs().download_by_path(photos_keys)

            await zip_writer.process(
                folder=folder,
                reference=item.get("reference", "SEM_REFERENCIA"),
                images_base64=photos_base64
            )

        await zip_writer.stream_to_cloud()
        return storage_key